import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from copy import copy


def consolidate_excel_files(directory, output_file):
    """
    Consolidates tables from Excel files in the specified directory into a new Excel file,
    copying the headers from the first file and then the rest of the data, with formatting.

    Args:
        directory (str): Path to the directory containing Excel files.
        output_file (str): Path to save the new consolidated Excel file.
    """
    # Create a new Excel workbook
    print("Creating Aggregated Data...")
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = "Consolidated Data"

    first_file = True  # Track if we are working with the first file
    start_row = 2  # Start from row 2 after headers

    # Loop through all Excel files in the directory
    for filename in os.listdir(directory):
        print(f"Copying {filename}...")
        if filename.endswith(".xlsx") and not filename.startswith("~$"):  # Ignore temp and other non-Excel files
            file_path = os.path.join(directory, filename)

            # Load the current workbook and select the active sheet
            current_wb = load_workbook(file_path)
            current_sheet = current_wb.active

            # Copy headers from the first file (row 3) to the new sheet's B1 row
            if first_file:
                copy_headers(current_sheet, new_sheet)
                first_file = False  # Turn off the first file flag

            # Get the title from cell A1
            title = current_sheet['A1'].value
            substring_to_remove = "Surveyor Skills Review (SSR)—"
            title = str(title).replace(substring_to_remove, "")

            # Loop through the table data starting at cell A4
            for row in current_sheet.iter_rows(min_row=4, min_col=1, max_col=current_sheet.max_column):
                # Write the title in the A column of the new sheet and copy its formatting
                title_cell = new_sheet.cell(row=start_row, column=1, value=title)
                # copy_cell_formatting(current_sheet['A1'], title_cell)

                # Copy table data and formatting starting in the B column of the new sheet
                for col_idx, source_cell in enumerate(row, start=2):
                    new_cell = new_sheet.cell(row=start_row, column=col_idx, value=source_cell.value)
                    copy_cell_formatting(source_cell, new_cell)

                start_row += 1  # Move to the next row for the next entry

    # Expanding column
    new_sheet.column_dimensions["A"].width = 45
    new_sheet.column_dimensions["B"].width = 50

    # Save the new workbook
    new_workbook.save(output_file)


def copy_headers(source_sheet, target_sheet):
    """
    Copies the contents and formatting of the header (third row) from the first file and pastes it into B1.

    Args:
        source_sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet from the first Excel file.
        target_sheet (openpyxl.worksheet.worksheet.Worksheet): The new sheet where headers will be pasted.
    """
    for col_idx, source_cell in enumerate(source_sheet[3], start=2):  # Row 3 of the first sheet contains headers
        new_cell = target_sheet.cell(row=1, column=col_idx,
                                     value=source_cell.value)  # Paste into row 1, starting from column B
        copy_cell_formatting(source_cell, new_cell)


def copy_cell_formatting(source_cell, target_cell):
    """
    Copies the formatting of a source cell to a target cell.

    Args:
        source_cell (openpyxl.cell.Cell): The source cell with the formatting to copy.
        target_cell (openpyxl.cell.Cell): The target cell where the formatting will be applied.
    """
    target_cell.font = copy(source_cell.font)
    target_cell.border = copy(source_cell.border)
    target_cell.fill = copy(source_cell.fill)
    target_cell.number_format = copy(source_cell.number_format)
    target_cell.protection = copy(source_cell.protection)
    target_cell.alignment = copy(source_cell.alignment)
